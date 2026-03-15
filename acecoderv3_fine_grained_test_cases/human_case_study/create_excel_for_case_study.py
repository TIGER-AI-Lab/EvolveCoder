#!/usr/bin/env python3
"""
Create an Excel workbook where each sheet corresponds to one Hugging Face dataset.
For each dataset:
  - randomly sample 10 examples
  - for each example randomly sample 10 tests
Formatting is optimized for human annotation:
  - Vertical layout per example:
      ID (single row)
      QUESTION (multi-line cell)
      TESTS (10 rows, one per test)
      (blank spacer rows)
  - Wide question column, wrapped text, frozen header space.

Requirements:
  pip install datasets openpyxl

Usage:
  1) Fill HF_URLS below
  2) python hf_to_excel.py
"""

from __future__ import annotations

import random
import re
from typing import List, Tuple

from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ----------------------------
# CONFIG
# ----------------------------
HF_URLS = [
    "chiruan/acecoderv3_round0",
    "chiruan/acecoderv3_round1"
]

OUTPUT_XLSX = "human_study.xlsx"
SPLIT_PREFERENCE = ["train", "validation", "test"]  # first available will be used
N_QUESTIONS = 10
N_TESTS = 10
RANDOM_SEED = 42  # set None for non-reproducible sampling

# Excel cell hard limit is 32,767 characters; truncate to avoid write errors.
EXCEL_CELL_CHAR_LIMIT = 32767
TRUNCATE_QUESTION_TO = 30000

# Layout
SPACER_ROWS_BETWEEN_EXAMPLES = 2


# ----------------------------
# HELPERS
# ----------------------------
def pick_split_name(ds_dict) -> str:
    keys = list(ds_dict.keys())
    for s in SPLIT_PREFERENCE:
        if s in ds_dict:
            return s
    return keys[0]


def load_one_dataset(url: str):
    ds = load_dataset(url)
    if hasattr(ds, "keys"):  # DatasetDict
        split = pick_split_name(ds)
        return ds[split], split
    return ds, "all"


def safe_sheet_name(name: str, existing: set[str]) -> str:
    """
    Excel sheet name rules:
      - max 31 chars
      - cannot contain: : \ / ? * [ ]
      - must be unique within workbook
    """
    name = re.sub(r"[:\\/?*\[\]]", "_", name).strip()
    if not name:
        name = "sheet"
    base = name[:31]

    if base not in existing:
        existing.add(base)
        return base

    i = 2
    while True:
        suffix = f"_{i}"
        trimmed = base[: 31 - len(suffix)]
        candidate = trimmed + suffix
        if candidate not in existing:
            existing.add(candidate)
            return candidate
        i += 1


def truncate_for_excel(s: str, limit: int) -> str:
    if s is None:
        return ""
    s = str(s)
    if len(s) <= limit:
        return s
    return s[:limit] + "\n\n[TRUNCATED]"


def sample_indices(n_total: int, n_sample: int, rng: random.Random) -> List[int]:
    if n_total <= 0:
        return []
    if n_total <= n_sample:
        return list(range(n_total))
    return rng.sample(range(n_total), n_sample)


def sample_tests(tests: List[str], n: int, rng: random.Random) -> List[str]:
    if not tests:
        return []
    if len(tests) <= n:
        return tests
    return rng.sample(tests, n)


def set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def style_cell(cell, wrap=True, valign="top", bold=False):
    cell.alignment = Alignment(wrap_text=wrap, vertical=valign)
    if bold:
        cell.font = Font(bold=True)


# ----------------------------
# MAIN
# ----------------------------
def main():
    if not HF_URLS:
        raise ValueError("HF_URLS is empty. Add your Hugging Face dataset IDs to HF_URLS.")

    rng = random.Random(RANDOM_SEED) if RANDOM_SEED is not None else random.Random()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    used_sheet_names: set[str] = set()

    # Styles
    label_font = Font(bold=True)
    label_fill = PatternFill("solid", fgColor="F2F2F2")
    mono_font = Font(name="Consolas")  # nice for tests
    thin_sep_fill = PatternFill("solid", fgColor="FFFFFF")

    for url in HF_URLS:
        ds, split = load_one_dataset(url)

        required_cols = {"id", "question", "tests"}
        missing = required_cols - set(ds.column_names)
        if missing:
            raise ValueError(
                f"Dataset '{url}' (split '{split}') missing columns: {sorted(missing)}. "
                f"Found columns: {ds.column_names}"
            )

        raw_sheet = f"{url.replace('/', '__')}__{split}"
        sheet_name = safe_sheet_name(raw_sheet, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)

        # Layout: two columns: A=Label, B=Content
        set_col_width(ws, 1, 14)   # label
        set_col_width(ws, 2, 120)  # content (wide for question/tests)

        # Optional: header note at top
        ws["A1"] = "Dataset"
        ws["B1"] = f"{url} (split: {split})"
        ws["A2"] = "Instructions"
        ws["B2"] = "Annotate top-to-bottom. Each example: ID → QUESTION → TESTS. Use spacer rows between examples."
        for r in (1, 2):
            ws[f"A{r}"].font = label_font
            ws[f"A{r}"].fill = label_fill
            ws[f"A{r}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws[f"B{r}"].alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A3"

        cur_row = 4  # start writing examples below header

        # Sample examples
        q_idxs = sample_indices(len(ds), N_QUESTIONS, rng)

        for ex_i, idx in enumerate(q_idxs, start=1):
            ex = ds[int(idx)]
            ex_id = truncate_for_excel(ex.get("id", ""), EXCEL_CELL_CHAR_LIMIT)
            question = truncate_for_excel(ex.get("question", ""), TRUNCATE_QUESTION_TO)

            tests = ex.get("tests", []) or []
            tests = [truncate_for_excel(t, EXCEL_CELL_CHAR_LIMIT) for t in tests]
            chosen_tests = sample_tests(tests, N_TESTS, rng)

            # --- ID row ---
            ws.cell(row=cur_row, column=1, value=f"ID ({ex_i})")
            ws.cell(row=cur_row, column=2, value=ex_id)
            for c in (1, 2):
                cell = ws.cell(row=cur_row, column=c)
                style_cell(cell, wrap=True, valign="top", bold=(c == 1))
                if c == 1:
                    cell.fill = label_fill
            cur_row += 1

            # --- QUESTION row ---
            ws.cell(row=cur_row, column=1, value="QUESTION")
            q_cell = ws.cell(row=cur_row, column=2, value=question)
            # Make question visually roomy
            ws.row_dimensions[cur_row].height = 220  # adjust as desired
            for c in (1, 2):
                cell = ws.cell(row=cur_row, column=c)
                style_cell(cell, wrap=True, valign="top", bold=(c == 1))
                if c == 1:
                    cell.fill = label_fill
            cur_row += 1

            # --- TESTS header row ---
            ws.cell(row=cur_row, column=1, value="TESTS")
            t_head = ws.cell(row=cur_row, column=2, value=f"{len(chosen_tests)} sampled test cases (one per row below)")
            for c in (1, 2):
                cell = ws.cell(row=cur_row, column=c)
                style_cell(cell, wrap=True, valign="top", bold=(c == 1))
                if c == 1:
                    cell.fill = label_fill
            cur_row += 1

            # --- Each test in its own row ---
            if not chosen_tests:
                ws.cell(row=cur_row, column=1, value="test_1")
                c2 = ws.cell(row=cur_row, column=2, value="")
                ws.cell(row=cur_row, column=1).fill = thin_sep_fill
                style_cell(ws.cell(row=cur_row, column=1), wrap=False, valign="top", bold=False)
                c2.font = mono_font
                style_cell(c2, wrap=True, valign="top", bold=False)
                cur_row += 1
            else:
                for j, t in enumerate(chosen_tests, start=1):
                    ws.cell(row=cur_row, column=1, value=f"test_{j}")
                    c2 = ws.cell(row=cur_row, column=2, value=t)
                    # Style
                    lcell = ws.cell(row=cur_row, column=1)
                    style_cell(lcell, wrap=False, valign="top", bold=False)
                    lcell.fill = thin_sep_fill

                    c2.font = mono_font
                    style_cell(c2, wrap=True, valign="top", bold=False)

                    # Optional: give tests some breathing room
                    ws.row_dimensions[cur_row].height = 60
                    cur_row += 1

            # Spacer rows between examples
            for _ in range(SPACER_ROWS_BETWEEN_EXAMPLES):
                cur_row += 1

    wb.save(OUTPUT_XLSX)
    print(f"Wrote Excel file: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
